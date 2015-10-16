import os
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from jllPortal import init_db
from array import *
import sqlite3

# configuration
DATA_FOLDER = os.path.join(os.getcwd(), 'data')
DATABASE = os.path.join(os.getcwd(), 'jll.db')

def connect_db():
	return sqlite3.connect(DATABASE)

def get_filepath(filename):
	return os.path.join(DATA_FOLDER, filename)

def insert_into(table, fields, values, data):
	cur.execute('insert into '+table+' '+fields+' values '+values, data)
	cur.execute('select max(id) FROM '+table)
	return cur.fetchone()[0]

def read_from(filename, sheet):
	wb = load_workbook(filename = get_filepath(filename), read_only=True)
	return wb[sheet]

def write_to(filename):
	return load_workbook(filename=get_filepath(filename))

def create(table, fields, values, worksheet, header):
	ids = array('i')
	for row in worksheet.rows:
		data = []
		if row[0].value != header:
			for cell in row:
				data.append(cell.value)
			ids.append(insert_into(table, fields, values, data))
	return ids

def write_column(worksheet, data, column, rowOffset):
	for i, entry in enumerate(data):
		worksheet.cell(row=i+rowOffset, column=column).value = entry

def create_clients():
	table = 'client'
	fields = '(name, imgFilename, industry)'
	values = '(?, ?, ?)'
	ws = read_from('Company.xlsx', 'Sheet1')
	return create(table, fields, values, ws, 'name')

def create_client_contacts(clientIds):
	table = 'clientContact'
	fields = '(firstName, lastName, email, phone, address, clientId)'
	values = '(?,?,?,?,?,?)'
	wb = write_to('CompanyContact.xlsx') 
	ws = wb['Sheet1']	
	write_column(ws, clientIds, 6, 2)
	wb.save(get_filepath('CompanyContact.xlsx'))
	return create(table, fields, values, ws, 'firstName')

def create_client_requirements(clientIds):
	table = 'clientRequirements'
	fields = '(RSF, budget, employees, market, clientId)'
	values = '(?,?,?,?,?)'
	wb = write_to('CompanyRequirements.xlsx')
	ws = wb['Sheet1']
	write_column(ws, clientIds, 5, 2)
	wb.save(get_filepath('CompanyRequirements.xlsx'))
	return create(table, fields, values, ws, 'RSF')

def create_property():
	table = 'property'
	fields = '(name, imagepath, address, description, size, rent, employees, market	)'
	values = '(?,?,?,?,?,?,?,?)'
	ws = read_from('Property.xlsx', 'Sheet1')
	return create(table, fields, values, ws, 'name')

def create_client_property(clientIds, propertyIds):
	table = 'client_property'
	fields = '(clientId, propertyId, comment, rating, showing)'
	values = '(?,?,?,?,?)'
	wb = write_to('CompanyProperty.xlsx')
	ws = wb['Sheet1']
	i = 0
	for clientId in clientIds:
		for propertyId in propertyIds:
			ws.cell(row=i+2, column=1).value = clientId
			ws.cell(row=i+2, column=2).value = propertyId
			ws.cell(row=i+2, column=5).value = 0
			i = i+1
	wb.save(get_filepath('CompanyProperty.xlsx'))
	return create(table, fields, values, ws, 'clientId')

conn = connect_db()
cur = conn.cursor()

init_db()



clientIds = create_clients()
create_client_contacts(clientIds)
propertyIds = create_property()
create_client_property(clientIds, propertyIds)

# Save (commit) the changes
conn.commit()

# We can also close the connection if we are done with it.
# Just be sure any changes have been committed or they will be lost.
conn.close()

def show_tables():
	conn = connect_db()
	cur = conn.cursor()
	print '****************************************'
	print '*****************Client*****************'
	print '****************************************'
	cur.execute('select * from client')
	for row in cur.fetchall():
		print row
	print '****************************************'
	print '*************Client Contact*************'
	print '****************************************'
	cur.execute('select * from clientContact')
	for row in cur.fetchall():
		print row
	print '****************************************'
	print '**********Client Requirements***********'
	print '****************************************'
	cur.execute('select * from clientRequirements')
	for row in cur.fetchall():
		print row
	print '****************************************'
	print '****************Property****************'
	print '****************************************'
	cur.execute('select * from property')
	for row in cur.fetchall():
		print row
	print '****************************************'
	print '************Client Property*************'
	print '****************************************'
	cur.execute('select * from client_property')
	for row in cur.fetchall():
		print row
	conn.close()